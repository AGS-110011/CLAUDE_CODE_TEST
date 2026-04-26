"""
Build the skeleton workbook: three sheets, headers, formats, validations,
ANALYSIS formulas, KEY_COMPARABLES lookups, charts.
Call create_workbook(data_end_row) → openpyxl.Workbook.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, ScatterChart, Series
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Colour constants (no leading #)
# ---------------------------------------------------------------------------
C_HEADER_FILL = "1F3864"   # navy header fill
C_HEADER_FONT = "FFFFFF"   # white
C_INPUT_FONT = "0000FF"    # blue — input cells
C_FORMULA_FONT = "000000"  # black — formula cells
C_XSHEET_FONT = "008000"   # green — cross-sheet links
C_KPI_FILL = "FFF2CC"      # yellow KPI
C_STRIPE = "F7F9FC"        # alternating row fill
C_SECTION = "D9E1F2"       # section header fill on ANALYSIS
C_SECTION_FONT = "1F3864"  # navy font for section headers

FONT_NAME = "Arial"

# ---------------------------------------------------------------------------
# INPUT_DATA column spec
# ---------------------------------------------------------------------------
INPUT_HEADERS = [
    "ID", "Type", "Link", "Address", "Distance (m)",
    "Price (€)", "Size (m2)", "Price per m2 (€)",
    "Bedrooms", "Bathrooms", "Floor", "Elevator", "Condition",
    "Year", "Terrace", "Parking", "Listing Date", "Days on Market", "Notes",
]
INPUT_WIDTHS = [6, 8, 32, 36, 12, 12, 10, 14, 11, 11, 8, 10, 16, 8, 9, 9, 14, 14, 40]

# Column index (1-based) shortcuts
COL_ID = 1
COL_TYPE = 2
COL_LINK = 3
COL_ADDRESS = 4
COL_DIST = 5
COL_PRICE = 6
COL_SIZE = 7
COL_PPM = 8
COL_BEDS = 9
COL_BATHS = 10
COL_FLOOR = 11
COL_ELEV = 12
COL_COND = 13
COL_YEAR = 14
COL_TERR = 15
COL_PARK = 16
COL_DATE = 17
COL_DOM = 18
COL_NOTES = 19


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _font(bold=False, color=C_FORMULA_FONT, size=10, italic=False):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size, italic=italic)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _header_style(cell, text: str) -> None:
    cell.value = text
    cell.font = _font(bold=True, color=C_HEADER_FONT, size=10)
    cell.fill = _fill(C_HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _section_header(cell, text: str) -> None:
    cell.value = text
    cell.font = _font(bold=True, color=C_SECTION_FONT, size=10)
    cell.fill = _fill(C_SECTION)


# ---------------------------------------------------------------------------
# Create INPUT_DATA sheet skeleton
# ---------------------------------------------------------------------------

def _build_input_sheet(wb: Workbook, data_end: int) -> None:
    ws = wb["INPUT_DATA"]
    ws.freeze_panes = "A2"

    # Headers row 1
    for col, (header, width) in enumerate(zip(INPUT_HEADERS, INPUT_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col)
        _header_style(cell, header)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 28

    # Number formats for data area
    FMT_DIST = '#,##0" m"'
    FMT_PRICE = '€#,##0'
    FMT_SIZE = '#,##0.0" m²"'
    FMT_PPM = '€#,##0'
    FMT_YEAR = '0'
    FMT_DATE = 'yyyy-mm-dd'
    FMT_DOM = '0" days"'

    col_fmts = {
        COL_DIST: FMT_DIST,
        COL_PRICE: FMT_PRICE,
        COL_SIZE: FMT_SIZE,
        COL_PPM: FMT_PPM,
        COL_YEAR: FMT_YEAR,
        COL_DATE: FMT_DATE,
        COL_DOM: FMT_DOM,
    }

    # Write row formulas and number formats for the full data area
    for r in range(2, data_end + 1):
        # Price per m²  (col H)
        c_ppm = ws.cell(row=r, column=COL_PPM)
        c_ppm.value = (
            f"=IFERROR(IF(AND(F{r}>0,G{r}>0),F{r}/G{r},\"\"),\"\")"
        )
        c_ppm.font = _font(color=C_FORMULA_FONT)

        # Days on market (col R)
        c_dom = ws.cell(row=r, column=COL_DOM)
        c_dom.value = f'=IFERROR(IF(Q{r}="","",TODAY()-Q{r}),"")'
        c_dom.font = _font(color=C_FORMULA_FONT)

        # Row stripe
        stripe = _fill(C_STRIPE) if r % 2 == 0 else None

        for col in range(1, 20):
            cell = ws.cell(row=r, column=col)
            if col in col_fmts:
                cell.number_format = col_fmts[col]
            if stripe and cell.value is None:
                cell.fill = stripe

    # Data validations
    _add_dv(ws, f"B2:B{data_end}", '"Sale,Rent"')
    _add_dv(ws, f"L2:L{data_end}", '"Yes,No"')
    _add_dv(ws, f"O2:O{data_end}", '"Yes,No"')
    _add_dv(ws, f"P2:P{data_end}", '"Yes,No"')
    _add_dv(ws, f"M2:M{data_end}", '"New,Renovated,To renovate"')


def _add_dv(ws, cell_range: str, formula1: str) -> None:
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True, showDropDown=False)
    dv.sqref = cell_range
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Create ANALYSIS sheet
# ---------------------------------------------------------------------------

def _build_analysis_sheet(wb: Workbook, data_end: int) -> None:
    ws = wb["ANALYSIS"]

    R = data_end
    PRICE = f"INPUT_DATA!$F$2:$F${R}"
    PPM = f"INPUT_DATA!$H$2:$H${R}"
    TYPE_ = f"INPUT_DATA!$B$2:$B${R}"
    COND = f"INPUT_DATA!$M$2:$M${R}"

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14

    # Row 1 — title (merged A1:D1, 14pt navy)
    ws.merge_cells("A1:D1")
    tc = ws["A1"]
    tc.value = "REAL ESTATE INVESTMENT — ANALYSIS DASHBOARD"
    tc.font = Font(name=FONT_NAME, bold=True, color=C_HEADER_FONT, size=14)
    tc.fill = _fill(C_HEADER_FILL)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Row 2 — subtitle
    ws["A2"].value = "All metrics auto-update from the INPUT_DATA sheet. Empty rows are ignored."
    ws["A2"].font = _font(italic=True, color="808080")
    ws.merge_cells("A2:D2")

    # --- Section A: Sales Analysis (rows 4–9) ---
    _section_merged(ws, "A4:D4", "A) SALES ANALYSIS  (price per m²)")
    _kv(ws, 5, "Average Price per m² (Sale)", f'=IFERROR(AVERAGEIFS({PPM},{TYPE_},"Sale",{PPM},">0"),0)', "€#,##0")
    _kv(ws, 6, "Median Price per m² (Sale)", f'=IFERROR(AGGREGATE(17,6,{PPM}/(({TYPE_}="Sale")*({PPM}>0)),2),0)', "€#,##0")
    _kv(ws, 7, "Min Price per m² (Sale)", f'=IFERROR(MINIFS({PPM},{TYPE_},"Sale",{PPM},">0"),0)', "€#,##0")
    _kv(ws, 8, "Max Price per m² (Sale)", f'=IFERROR(MAXIFS({PPM},{TYPE_},"Sale",{PPM},">0"),0)', "€#,##0")
    _kv(ws, 9, "Count (Sale listings)", f'=COUNTIFS({TYPE_},"Sale",{PRICE},">0")', "0")

    # --- Section B: Rent Analysis (rows 11–15) ---
    _section_merged(ws, "A11:D11", "B) RENT ANALYSIS  (monthly €)")
    _kv(ws, 12, "Avg Monthly Rent", f'=IFERROR(AVERAGEIFS({PRICE},{TYPE_},"Rent",{PRICE},">0"),0)', "€#,##0")
    _kv(ws, 13, "Avg Rent per m²", f'=IFERROR(AVERAGEIFS({PPM},{TYPE_},"Rent",{PPM},">0"),0)', "€#,##0")
    _kv(ws, 14, "Median Rent per m²", f'=IFERROR(AGGREGATE(17,6,{PPM}/(({TYPE_}="Rent")*({PPM}>0)),2),0)', "€#,##0")
    _kv(ws, 15, "Count (Rent listings)", f'=COUNTIFS({TYPE_},"Rent",{PRICE},">0")', "0")

    # --- Section C: Yield Estimation (rows 17–20) ---
    _section_merged(ws, "A17:D17", "C) YIELD ESTIMATION")
    _kv(ws, 18, "Avg Sale Price (whole property)", f'=IFERROR(AVERAGEIFS({PRICE},{TYPE_},"Sale",{PRICE},">0"),0)', "€#,##0")
    _kv(ws, 19, "Annualised Rent (B12×12)", "=B12*12", "€#,##0")
    kpi_cell = ws.cell(row=20, column=2)
    kpi_cell.value = "=IFERROR(IF(B18>0,B19/B18,0),0)"
    kpi_cell.number_format = "0.00%"
    kpi_cell.fill = _fill(C_KPI_FILL)
    kpi_cell.font = _font(bold=True)
    ws.cell(row=20, column=1).value = "Gross Rental Yield"
    ws.cell(row=20, column=1).font = _font()
    ws.cell(row=20, column=3).value = "← headline KPI"
    ws.cell(row=20, column=3).font = _font(bold=True, color=C_HEADER_FILL)

    # --- Section D: Market Segmentation (rows 22–26) ---
    _section_merged(ws, "A22:D22", "D) MARKET SEGMENTATION  (Sale listings)")
    # sub-headers row 23
    for col, hdr in enumerate(["Condition", "Avg €/m²", "Median €/m²", "# Listings"], start=1):
        c = ws.cell(row=23, column=col)
        c.value = hdr
        c.font = _font(bold=True, color=C_HEADER_FONT)
        c.fill = _fill(C_HEADER_FILL)
        c.alignment = Alignment(horizontal="center")

    for row, cond in [(24, "New"), (25, "Renovated"), (26, "To renovate")]:
        ws.cell(row=row, column=1).value = cond
        ws.cell(row=row, column=2).value = (
            f'=IFERROR(AVERAGEIFS({PPM},{TYPE_},"Sale",{COND},"{cond}",{PPM},">0"),0)'
        )
        ws.cell(row=row, column=2).number_format = "€#,##0"
        ws.cell(row=row, column=3).value = (
            f'=IFERROR(AGGREGATE(17,6,{PPM}/(({TYPE_}="Sale")*({COND}="{cond}")*({PPM}>0)),2),0)'
        )
        ws.cell(row=row, column=3).number_format = "€#,##0"
        ws.cell(row=row, column=4).value = (
            f'=COUNTIFS({TYPE_},"Sale",{COND},"{cond}",{PRICE},">0")'
        )
        ws.cell(row=row, column=4).number_format = "0"
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = _font()

    # --- Section E: Calculation Notes (rows 28–31) ---
    _section_merged(ws, "A28:D28", "E) CALCULATION NOTES")
    notes_text = [
        "Medians use AGGREGATE(17,…) which ignores errors and blank cells automatically.",
        "All averages use AVERAGEIFS with >0 guard so empty rows are excluded from statistics.",
        "To extend the data area beyond 200 rows, add data to INPUT_DATA and the formulas update automatically.",
    ]
    for i, txt in enumerate(notes_text):
        cell = ws.cell(row=29 + i, column=1)
        cell.value = txt
        cell.font = _font(italic=True, color="808080")
        ws.merge_cells(f"A{29+i}:D{29+i}")

    # --- Charts ---
    ws_input = wb["INPUT_DATA"]

    # Bar chart: Distribution of price per m²  (anchor F4)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Distribution of price per m² (all listings)"
    bar.y_axis.title = "Price per m² (€)"
    bar.x_axis.title = "Listing ID"
    bar.width = 18
    bar.height = 9
    bar.legend = None

    bar_data = Reference(ws_input, min_col=COL_PPM, max_col=COL_PPM, min_row=1, max_row=R)
    bar.add_data(bar_data, titles_from_data=True)
    bar_cats = Reference(ws_input, min_col=COL_ID, max_col=COL_ID, min_row=2, max_row=R)
    bar.set_categories(bar_cats)
    ws.add_chart(bar, "F4")

    # Scatter chart: Price vs Size  (anchor F22)
    scatter = ScatterChart()
    scatter.title = "Price vs Size (all listings)"
    scatter.x_axis.title = "Size (m²)"
    scatter.y_axis.title = "Price (€)"
    scatter.width = 18
    scatter.height = 9
    scatter.legend = None

    xvals = Reference(ws_input, min_col=COL_SIZE, max_col=COL_SIZE, min_row=2, max_row=R)
    yvals = Reference(ws_input, min_col=COL_PRICE, max_col=COL_PRICE, min_row=2, max_row=R)
    s = Series(yvals, xvals, title="Properties")
    scatter.series.append(s)
    ws.add_chart(scatter, "F22")


def _section_merged(ws, cell_range: str, text: str) -> None:
    ws.merge_cells(cell_range)
    start_cell = cell_range.split(":")[0]
    c = ws[start_cell]
    _section_header(c, text)


def _kv(ws, row: int, label: str, formula: str, fmt: str) -> None:
    ws.cell(row=row, column=1).value = label
    ws.cell(row=row, column=1).font = _font()
    val_cell = ws.cell(row=row, column=2)
    val_cell.value = formula
    val_cell.number_format = fmt
    val_cell.font = _font(color=C_FORMULA_FONT)


# ---------------------------------------------------------------------------
# Create KEY_COMPARABLES sheet
# ---------------------------------------------------------------------------

def _build_comparables_sheet(wb: Workbook, data_end: int) -> None:
    ws = wb["KEY_COMPARABLES"]
    R = data_end

    COMP_HEADERS = [
        "ID", "Type", "Price", "Size", "Price per m²",
        "Condition", "Distance", "Comment (why it is relevant)",
    ]
    COMP_WIDTHS = [8, 10, 14, 10, 14, 16, 12, 50]

    # Row 1 — title
    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value = "KEY COMPARABLES"
    tc.font = Font(name=FONT_NAME, bold=True, color=C_HEADER_FONT, size=14)
    tc.fill = _fill(C_HEADER_FILL)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Row 2 — instructions
    ws.merge_cells("A2:H2")
    ws["A2"].value = (
        "Enter an ID from INPUT_DATA in column A to auto-fill the row. "
        "Column H is for manual comments."
    )
    ws["A2"].font = _font(italic=True, color="808080")

    # Row 3 — blank separator

    # Row 4 — headers
    for col, (hdr, width) in enumerate(zip(COMP_HEADERS, COMP_WIDTHS), start=1):
        c = ws.cell(row=4, column=col)
        _header_style(c, hdr)
        ws.column_dimensions[get_column_letter(col)].width = width

    # Rows 5–24 — lookup rows
    # Mapping: col B→Type($B), C→Price($F), D→Size($G), E→PPM($H), F→Cond($M), G→Dist($E)
    src_cols = {
        2: "B",   # Type
        3: "F",   # Price
        4: "G",   # Size
        5: "H",   # Price per m²
        6: "M",   # Condition
        7: "E",   # Distance
    }
    fmts = {
        3: "€#,##0",
        4: '#,##0.0" m²"',
        5: "€#,##0",
        7: '#,##0" m"',
    }

    for row in range(5, 25):
        for col, src_col in src_cols.items():
            formula = (
                f'=IF($A{row}="","",IFERROR('
                f'INDEX(INPUT_DATA!${src_col}$2:${src_col}${R},'
                f'MATCH($A{row},INPUT_DATA!$A$2:$A${R},0)),"not found"))'
            )
            cell = ws.cell(row=row, column=col)
            cell.value = formula
            cell.font = _font(color=C_XSHEET_FONT)
            if col in fmts:
                cell.number_format = fmts[col]
        # ID column style (input)
        id_cell = ws.cell(row=row, column=1)
        id_cell.font = _font(color=C_INPUT_FONT)
        # Comment column — manual input
        ws.cell(row=row, column=8).font = _font(color=C_INPUT_FONT)


# ---------------------------------------------------------------------------
# Public factory
# ---------------------------------------------------------------------------

def create_workbook(data_end_row: int) -> Workbook:
    """
    Build and return an openpyxl Workbook with all three sheets populated
    with structure, formatting, formulas, and charts. Data rows are empty.
    """
    wb = Workbook()

    # Rename default sheet and add the other two
    wb.active.title = "INPUT_DATA"
    wb.create_sheet("ANALYSIS")
    wb.create_sheet("KEY_COMPARABLES")

    _build_input_sheet(wb, data_end_row)
    _build_analysis_sheet(wb, data_end_row)
    _build_comparables_sheet(wb, data_end_row)

    return wb
