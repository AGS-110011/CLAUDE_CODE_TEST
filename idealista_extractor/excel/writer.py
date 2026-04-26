"""
Write extracted listings into the workbook and run LibreOffice recalculation.
"""
from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rich.console import Console

from ..models import Listing
from .template import (
    C_FORMULA_FONT,
    C_INPUT_FONT,
    C_STRIPE,
    COL_ADDRESS,
    COL_BATHS,
    COL_BEDS,
    COL_COND,
    COL_DATE,
    COL_DIST,
    COL_ELEV,
    COL_FLOOR,
    COL_ID,
    COL_LINK,
    COL_NOTES,
    COL_PARK,
    COL_PRICE,
    COL_SIZE,
    COL_TERR,
    COL_TYPE,
    COL_YEAR,
    FONT_NAME,
    create_workbook,
    _fill,
)

console = Console()

# Columns written as raw data (others are formulas set by template)
_DATA_COLS = {
    COL_ID, COL_TYPE, COL_LINK, COL_ADDRESS, COL_DIST, COL_PRICE,
    COL_SIZE, COL_BEDS, COL_BATHS, COL_FLOOR, COL_ELEV, COL_COND,
    COL_YEAR, COL_TERR, COL_PARK, COL_DATE, COL_NOTES,
}

FMT_DIST = '#,##0" m"'
FMT_PRICE = "€#,##0"
FMT_SIZE = '#,##0.0" m²"'
FMT_PPM = "€#,##0"
FMT_YEAR = "0"
FMT_DATE = "yyyy-mm-dd"
FMT_DOM = '0" days"'

_COL_FMTS = {
    COL_DIST: FMT_DIST,
    COL_PRICE: FMT_PRICE,
    COL_SIZE: FMT_SIZE,
    COL_YEAR: FMT_YEAR,
    COL_DATE: FMT_DATE,
}


def _compute_data_end(n_listings: int) -> int:
    if n_listings <= 200:
        return 201
    return n_listings + 51


def _write_row(ws, row: int, listing: Listing) -> None:
    stripe = _fill(C_STRIPE) if row % 2 == 0 else None

    values = {
        COL_ID: listing.id,
        COL_TYPE: listing.type,
        COL_LINK: str(listing.link),
        COL_ADDRESS: listing.address,
        COL_DIST: listing.distance_m,
        COL_PRICE: listing.price_eur,
        COL_SIZE: listing.size_m2,
        COL_BEDS: listing.bedrooms,
        COL_BATHS: listing.bathrooms,
        COL_FLOOR: listing.floor,
        COL_ELEV: listing.elevator,
        COL_COND: listing.condition,
        COL_YEAR: listing.year,
        COL_TERR: listing.terrace,
        COL_PARK: listing.parking,
        COL_DATE: listing.listing_date,
        COL_NOTES: listing.notes,
    }

    for col, val in values.items():
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.font = Font(name=FONT_NAME, color=C_INPUT_FONT, size=10)
        if col in _COL_FMTS:
            cell.number_format = _COL_FMTS[col]
        if stripe:
            cell.fill = stripe

    # Link as hyperlink
    link_cell = ws.cell(row=row, column=COL_LINK)
    link_url = str(listing.link)
    if link_url.startswith("http"):
        link_cell.hyperlink = link_url
        link_cell.value = link_url
        link_cell.font = Font(
            name=FONT_NAME, color="0563C1", underline="single", size=10
        )


def write_workbook(
    sale_listings: list[Listing],
    rent_listings: list[Listing],
    out_path: str,
) -> None:
    """
    Create the workbook, fill all sheets, save, run LibreOffice recalc,
    then verify KPI cells contain numbers.
    """
    total = len(sale_listings) + len(rent_listings)
    data_end = _compute_data_end(total)

    console.print(f"[bold]Building workbook[/bold]: {total} listings, data area rows 2–{data_end}")

    wb = create_workbook(data_end)
    ws = wb["INPUT_DATA"]

    # Assign IDs: sale 1..Ns, rent Ns+1..Ns+Nr
    row = 2
    for i, lst in enumerate(sale_listings, start=1):
        lst.id = i
        _write_row(ws, row, lst)
        row += 1

    rent_start = len(sale_listings) + 1
    for i, lst in enumerate(rent_listings, start=rent_start):
        lst.id = i
        _write_row(ws, row, lst)
        row += 1

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    console.print(f"[green]Workbook saved → {out}[/green]")

    # LibreOffice recalculation pass
    _recalc_with_libreoffice(out)

    # Verify KPI cells
    _verify_kpis(out)


# ---------------------------------------------------------------------------
# LibreOffice recalculation
# ---------------------------------------------------------------------------

def _find_soffice() -> Optional[str]:
    for candidate in [
        "soffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/usr/bin/soffice",
        "/usr/lib/libreoffice/program/soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]:
        try:
            result = subprocess.run(
                [candidate, "--version"],
                capture_output=True,
                timeout=10,
            )
            if result.returncode == 0:
                return candidate
        except (FileNotFoundError, subprocess.TimeoutExpired, PermissionError):
            continue
    return None


def _recalc_with_libreoffice(out: Path) -> None:
    soffice = _find_soffice()
    if not soffice:
        console.print(
            "[yellow]LibreOffice not found — skipping recalculation pass. "
            "Dashboard cells may show 0 until the file is opened in Excel.[/yellow]"
        )
        return

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        console.print("[cyan]Running LibreOffice headless recalculation…[/cyan]")
        try:
            subprocess.run(
                [
                    soffice,
                    "--headless",
                    "--calc",
                    "--convert-to", "xlsx",
                    str(out),
                    "--outdir", tmp,
                ],
                check=True,
                timeout=120,
                capture_output=True,
            )
            recalced = tmp_path / out.name
            if recalced.exists():
                shutil.move(str(recalced), str(out))
                console.print(f"[green]LibreOffice recalculation complete → {out}[/green]")
            else:
                console.print("[yellow]LibreOffice did not produce output; using original file.[/yellow]")
        except subprocess.CalledProcessError as exc:
            console.print(f"[yellow]LibreOffice error (non-fatal): {exc}[/yellow]")
        except subprocess.TimeoutExpired:
            console.print("[yellow]LibreOffice timed out; using original file.[/yellow]")


# ---------------------------------------------------------------------------
# KPI verification
# ---------------------------------------------------------------------------

def _verify_kpis(out: Path) -> None:
    try:
        wb = load_workbook(str(out), data_only=True)
        ws = wb["ANALYSIS"]

        kpis = {
            "Avg sale €/m² (B5)": ws["B5"].value,
            "Avg monthly rent (B12)": ws["B12"].value,
            "Gross yield (B20)": ws["B20"].value,
        }

        all_ok = True
        for name, val in kpis.items():
            if val is None or val == "" or (isinstance(val, str) and val.startswith("#")):
                console.print(f"[red]KPI BLANK/ERROR: {name} = {val!r}[/red]")
                all_ok = False
            else:
                console.print(f"[green]KPI OK: {name} = {val}[/green]")

        if not all_ok:
            console.print(
                "[bold red]One or more KPI cells are blank. "
                "Open the file in Excel and press Ctrl+Alt+F9 to force recalculation.[/bold red]"
            )
            sys.exit(1)
    except Exception as exc:
        console.print(f"[yellow]KPI verification skipped: {exc}[/yellow]")
