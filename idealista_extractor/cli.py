"""CLI entry-point for idealista-extractor."""
from __future__ import annotations

import asyncio
import re
import sys
from typing import Optional

import click
from rich.console import Console
from rich.progress import BarColumn, Progress, SpinnerColumn, TaskProgressColumn, TextColumn
from rich.table import Table

from .config import Config
from .models import Listing
from .scraper.browser import BrowserSession
from .scraper.listing import extract_listing
from .scraper.search import paginate_search

console = Console()

_URL_RE = re.compile(
    r"^https://www\.idealista\.(com|es)(/en)?/areas/(venta|alquiler)-viviendas/.*\?shape=.+"
)


def _validate_url(ctx, param, value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    if not _URL_RE.match(value):
        raise click.BadParameter(
            "URL must match: https://www.idealista.(com|es)/areas/(venta|alquiler)-viviendas/…?shape=…"
        )
    return value


@click.command("idealista-extractor")
@click.option("--sale-url", default=None, callback=_validate_url, help="Idealista sale areas URL with shape param.")
@click.option("--rent-url", default=None, callback=_validate_url, help="Idealista rent areas URL with shape param.")
@click.option("--target-lat", default=None, type=float, help="Reference property latitude for distance calc.")
@click.option("--target-lon", default=None, type=float, help="Reference property longitude for distance calc.")
@click.option("--out", default="./Idealista_Property_Analysis.xlsx", show_default=True, help="Output .xlsx path.")
@click.option("--max-per-type", default=100, show_default=True, type=click.IntRange(1, 200), help="Max listings per URL.")
@click.option("--delay-min", default=4.0, show_default=True, type=float, help="Min delay between requests (s).")
@click.option("--delay-max", default=8.0, show_default=True, type=float, help="Max delay between requests (s).")
@click.option("--headful", is_flag=True, default=False, help="Show browser window (for first-run cookie priming).")
@click.option("--session-file", default="./.idealista_session.json", show_default=True, help="Session state path.")
def main(
    sale_url: Optional[str],
    rent_url: Optional[str],
    target_lat: Optional[float],
    target_lon: Optional[float],
    out: str,
    max_per_type: int,
    delay_min: float,
    delay_max: float,
    headful: bool,
    session_file: str,
) -> None:
    """
    Scrape Idealista area-search listings and write an Excel investment workbook.

    At least one of --sale-url / --rent-url is required.
    """
    if not sale_url and not rent_url:
        raise click.UsageError("Provide at least one of --sale-url / --rent-url.")

    cfg = Config(
        sale_url=sale_url,
        rent_url=rent_url,
        target_lat=target_lat,
        target_lon=target_lon,
        out=out,
        max_per_type=max_per_type,
        delay_min=delay_min,
        delay_max=delay_max,
        headful=headful,
        session_file=session_file,
    )

    asyncio.run(_run(cfg))


async def _run(cfg: Config) -> None:
    from .excel.writer import write_workbook

    sale_urls: list[str] = []
    rent_urls: list[str] = []
    sale_listings: list[Listing] = []
    rent_listings: list[Listing] = []

    async with BrowserSession(
        session_file=cfg.session_file,
        headful=cfg.headful,
        debug_dir=cfg.debug_dir,
    ) as session:

        # ---- Phase 1: collect listing URLs ----
        if cfg.sale_url:
            console.rule("[bold cyan]Sale: collecting listing URLs[/bold cyan]")
            try:
                sale_urls = await paginate_search(
                    session, cfg.sale_url, "Sale", cfg.max_per_type,
                    cfg.delay_min, cfg.delay_max,
                )
            except RuntimeError as exc:
                console.print(f"[bold red]{exc}[/bold red]")
                sys.exit(1)

        if cfg.rent_url:
            console.rule("[bold cyan]Rent: collecting listing URLs[/bold cyan]")
            try:
                rent_urls = await paginate_search(
                    session, cfg.rent_url, "Rent", cfg.max_per_type,
                    cfg.delay_min, cfg.delay_max,
                )
            except RuntimeError as exc:
                console.print(f"[bold red]{exc}[/bold red]")
                sys.exit(1)

        n_sale = len(sale_urls)
        n_rent = len(rent_urls)
        est_mins = round((n_sale + n_rent) * (cfg.delay_min + cfg.delay_max) / 2 / 60)
        console.rule()
        console.print(
            f"[bold green]Found {n_sale} sale + {n_rent} rent listings matching shape. "
            f"Capped at {cfg.max_per_type} + {cfg.max_per_type}. "
            f"Est. runtime ~{est_mins} min at {(cfg.delay_min+cfg.delay_max)/2:.0f}s/page.[/bold green]"
        )

        # ---- Phase 2: extract individual listings ----
        if sale_urls:
            console.rule("[bold cyan]Extracting sale listings[/bold cyan]")
            sale_listings = await _extract_batch(
                session, sale_urls, "Sale", cfg, label="Listings (sale)"
            )

        if rent_urls:
            console.rule("[bold cyan]Extracting rent listings[/bold cyan]")
            rent_listings = await _extract_batch(
                session, rent_urls, "Rent", cfg, label="Listings (rent)"
            )

    # ---- Phase 3: write workbook ----
    console.rule("[bold cyan]Writing Excel workbook[/bold cyan]")
    try:
        write_workbook(sale_listings, rent_listings, cfg.out)
    except SystemExit:
        raise
    except Exception as exc:
        console.print(f"[bold red]Failed to write workbook: {exc}[/bold red]")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    # ---- Phase 4: print summary ----
    _print_summary(sale_listings, rent_listings, cfg.out)


async def _extract_batch(
    session: BrowserSession,
    urls: list[str],
    listing_type: str,
    cfg: Config,
    label: str,
) -> list[Listing]:
    results: list[Listing] = []
    failed = 0

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        console=console,
        transient=False,
    ) as progress:
        task = progress.add_task(f"[cyan]{label}[/cyan]", total=len(urls))

        for url in urls:
            try:
                raw = await extract_listing(
                    session, url, listing_type,
                    cfg.delay_min, cfg.delay_max,
                    cfg.target_lat, cfg.target_lon,
                )
                results.append(Listing(**raw))
            except Exception as exc:
                console.print(f"  [red]Warning: {url} → {exc}[/red]")
                failed += 1
            progress.advance(task)

    console.print(
        f"  [green]{len(results)} extracted[/green], "
        f"[red]{failed} failed[/red] for {listing_type.lower()} listings."
    )
    return results


def _print_summary(
    sale: list[Listing],
    rent: list[Listing],
    out_path: str,
) -> None:
    from openpyxl import load_workbook

    table = Table(title="Extraction Summary", show_header=True, header_style="bold magenta")
    table.add_column("Metric", style="cyan", no_wrap=True)
    table.add_column("Value", style="white")

    table.add_row("Sale listings extracted", str(len(sale)))
    table.add_row("Rent listings extracted", str(len(rent)))
    table.add_row("Total", str(len(sale) + len(rent)))
    table.add_row("Output file", out_path)

    # Try to read KPIs from the saved workbook
    try:
        wb = load_workbook(out_path, data_only=True)
        ws = wb["ANALYSIS"]
        avg_sale_ppm = ws["B5"].value
        avg_rent = ws["B12"].value
        gross_yield = ws["B20"].value

        if avg_sale_ppm is not None:
            table.add_row("Avg sale €/m² (dashboard)", f"€{avg_sale_ppm:,.0f}" if isinstance(avg_sale_ppm, (int, float)) else str(avg_sale_ppm))
        if avg_rent is not None:
            table.add_row("Avg monthly rent (dashboard)", f"€{avg_rent:,.0f}" if isinstance(avg_rent, (int, float)) else str(avg_rent))
        if gross_yield is not None:
            table.add_row(
                "Gross rental yield (KPI)",
                f"{gross_yield*100:.2f}%" if isinstance(gross_yield, float) else str(gross_yield),
                style="bold green",
            )
    except Exception:
        pass

    console.print()
    console.print(table)
    console.print(f"\n[bold green]Done. Open {out_path} in Excel or LibreOffice.[/bold green]")
